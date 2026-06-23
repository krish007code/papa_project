import glob
import io
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


# ── helpers ───────────────────────────────────────────────────────────────────

HAIR = Side(border_style="hair", color="000000")

def _hair_border(top: bool = True, bottom: bool = True) -> Border:
    return Border(
        left=HAIR,  right=HAIR,
        top=HAIR    if top    else Side(),
        bottom=HAIR if bottom else Side(),
    )

def _age_bucket(raw_str: str, now: datetime) -> str:
    try:
        dt   = datetime.strptime(str(raw_str).strip(), "%d-%m-%Y %H:%M:%S")
        days = (now - dt).total_seconds() / 86400
        if days < 1:
            return "Less than 1 Day"
        elif days <= 3:
            return "More than 1 Day"
        elif days <= 7:
            return "More than 3 Days"
        else:
            return "More than 7 Days"
    except Exception:
        return "Unknown"

def _find_file(base: Path, folder: str, pattern: str) -> Path:
    matches = glob.glob(str(base / folder / pattern))
    if not matches:
        raise FileNotFoundError(f"No file matching '{pattern}' in {base / folder}/")
    return Path(sorted(matches)[-1])


# ── core ──────────────────────────────────────────────────────────────────────

def generate_report(
    permanent_path: str | Path | io.BytesIO,
    monthly_path:   str | Path | io.BytesIO,
    daily_path:     str | Path | io.BytesIO,
    daily_path2:    str | Path | io.BytesIO,
    now:            datetime | None = None,
    generated_by:   str = "bsupwag2",
) -> bytes:
    now = now or datetime.now()

    # ── 1. BA / OA mapping ────────────────────────────────────────────────────
    ba_oa = pd.read_excel(permanent_path, dtype=str)
    ba_oa.columns = ba_oa.columns.str.strip()
    ba_oa_map = (
        ba_oa[["BA", "OA", "BLOCK"]]
        .copy()
        .assign(BLOCK=lambda d: d["BLOCK"].str.strip().str.upper())
        .drop_duplicates(subset=["BLOCK"])
    )
    # ── 2. Daily report ───────────────────────────────────────────────────────
    report = pd.read_excel(daily_path, skiprows=3, dtype=str)
    report.columns = report.columns.str.strip()
    report.dropna(how="all", inplace=True)
    report.reset_index(drop=True, inplace=True)
    report["_BLOCK_KEY"] = report["BLOCK"].str.strip().str.upper()

    # ── 3. Merge BA / OA before DISTRICT ─────────────────────────────────────
    merged = report.merge(
        ba_oa_map.rename(columns={"BLOCK": "_BLOCK_KEY"}),
        on="_BLOCK_KEY", how="left",
    )
    merged.drop(columns=["_BLOCK_KEY"], inplace=True)

    cols = list(merged.columns)
    for c in ["BA", "OA"]:
        if c in cols:
            cols.remove(c)
    di   = cols.index("DISTRICT")
    cols = cols[:di] + ["BA", "OA"] + cols[di:]
    merged = merged[cols]

    # ── 4. Age bucket ─────────────────────────────────────────────────────────
    merged["No of Days"] = merged["STATE CHANGE TIME"].apply(
        lambda v: _age_bucket(v, now)
    )

    # ── 5. AMC GP counts from monthly file ───────────────────────────────────
    xl_monthly = pd.ExcelFile(monthly_path)
    olt_sheets = [s for s in xl_monthly.sheet_names if s.lower().startswith("olt")]
    if not olt_sheets:
        raise ValueError("No sheet starting with 'OLT' found in the monthly file.")
    olt_sheet = sorted(olt_sheets)[-1]

    olt_raw = pd.read_excel(monthly_path, sheet_name=olt_sheet, skiprows=1, header=0)
    olt_raw.columns = olt_raw.columns.str.strip()

    amc_gp = (
        olt_raw[["OLT IP", "AMC GP Count"]]
        .dropna(subset=["OLT IP"])
        .copy()
        .assign(**{"OLT IP": lambda d: d["OLT IP"].astype(str).str.strip()})
    )
    amc_gp["AMC GP Count"] = pd.to_numeric(amc_gp["AMC GP Count"], errors="coerce")
    amc_gp = amc_gp.groupby("OLT IP", as_index=False)["AMC GP Count"].sum()
    amc_gp.rename(columns={"AMC GP Count": "No of AMC GPs"}, inplace=True)

    # ── 6. Join AMC GP on BLOCK NODE IP ──────────────────────────────────────
    merged["_OLT_KEY"] = merged["BLOCK NODE IP"].astype(str).str.strip()
    final = merged.merge(
        amc_gp.rename(columns={"OLT IP": "_OLT_KEY"}),
        on="_OLT_KEY", how="left",
    )
    final.drop(columns=["_OLT_KEY"], inplace=True)
    final["No of AMC GPs"] = (
        pd.to_numeric(final["No of AMC GPs"], errors="coerce").fillna(0).astype(int)
    )

    # ── 6.5 Count GPs with Unknown Previous UP status from daily2 file ───────
    daily2_raw = pd.read_excel(daily_path2, skiprows=3, dtype=str)
    daily2_raw.columns = daily2_raw.columns.str.strip()
    
    # Filter for UNKNOWN PREV UP status
    unknown_gps = (
        daily2_raw[daily2_raw["GP STATUS"].str.strip().str.upper() == "UNKNOWN PREV UP"]
        .copy()
    )
    
    # Count by OLT IP
    if not unknown_gps.empty:
        unknown_gp_count = (
            unknown_gps.groupby("OLT IP", as_index=False)
            .size()
            .rename(columns={"size": "No of GPs Unknown previously UP"})
        )
        unknown_gp_count["OLT IP"] = unknown_gp_count["OLT IP"].astype(str).str.strip()
    else:
        unknown_gp_count = pd.DataFrame(columns=["OLT IP", "No of GPs Unknown previously UP"])
    
    # Join unknown GP count on BLOCK NODE IP
    final["_OLT_KEY"] = final["BLOCK NODE IP"].astype(str).str.strip()
    final = final.merge(
        unknown_gp_count.rename(columns={"OLT IP": "_OLT_KEY"}),
        on="_OLT_KEY", how="left",
    )
    final.drop(columns=["_OLT_KEY"], inplace=True)
    final["No of GPs Unknown previously UP"] = (
        pd.to_numeric(final["No of GPs Unknown previously UP"], errors="coerce").fillna(0).astype(int)
    )

    # ── 7. Select and order final columns ────────────────────────────────────
    OUTPUT_COLS = [
        "BA", "OA", "DISTRICT", "BLOCK", "BLOCK NODE LOCATION",
        "BLOCK NODE IP", "ALARM REASON", "VENDOR",
        "STATE CHANGE TIME", "PHASE",
        "No of Days", "No of AMC GPs", "No of GPs Unknown previously UP",
    ]
    final = final[[c for c in OUTPUT_COLS if c in final.columns]]
    final.reset_index(drop=True, inplace=True)
    N_ROWS = len(final)

    # ── 8. Build Excel workbook ───────────────────────────────────────────────
    today_str    = now.strftime("%d/%m/%Y")
    generated_at = now.strftime("%d-%m-%Y %H:%M:%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "OLT Down Report"

    # Row 1 — title
    title_text = (
        f" Total NON_OPERATIONAL Report  STATE NAME : UTTAR PRADESH WEST"
        f"  Date : {today_str} ,Time : 10:00  and  Phase : BHARATNET"
    )
    ws.merge_cells("A1:N1")
    ws["A1"] = title_text
    ws["A1"].font      = Font(name="Courier New", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Row 2 — record count
    ws.merge_cells("A2:N2")
    ws["A2"] = f"Showing Records 1 to {N_ROWS}   , Report Generated at : {generated_at}"
    ws["A2"].font      = Font(name="Times New Roman", size=12)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15.75

    # Row 3 — blank
    ws.merge_cells("A3:N3")
    ws["A3"] = ""
    ws.row_dimensions[3].height = 15.75

    # Row 4 — headers
    HEADERS = [
        "Sr.No", "BA", "OA", "DISTRICT", "BLOCK", "BLOCK NODE LOCATION",
        "BLOCK NODE IP", "ALARM REASON", "VENDOR", "STATE CHANGE TIME",
        "PHASE", "No of Days", "No of AMC GPs", "No of GPs Unknown previously UP",
    ]
    ws.row_dimensions[4].height = 131.25
    for col_idx, hdr in enumerate(HEADERS, start=1):
        cell           = ws.cell(row=4, column=col_idx, value=hdr)
        cell.font      = Font(name="Times New Roman", bold=True, size=14)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = _hair_border(top=col_idx <= 11, bottom=col_idx <= 11)

    # Rows 5+ — data
    DATA_FONT  = Font(name="Times New Roman", size=12)
    EXTRA_FONT = Font(name="Calibri", size=11)

    for row_idx, (_, row) in enumerate(final.iterrows(), start=5):
        sr_no = row_idx - 4

        c        = ws.cell(row=row_idx, column=1, value=sr_no)
        c.font   = DATA_FONT
        c.border = _hair_border()

        data_vals = [
            row.get("BA", ""),                  row.get("OA", ""),
            row.get("DISTRICT", ""),             row.get("BLOCK", ""),
            row.get("BLOCK NODE LOCATION", ""),  row.get("BLOCK NODE IP", ""),
            row.get("ALARM REASON", ""),         row.get("VENDOR", ""),
            row.get("STATE CHANGE TIME", ""),    row.get("PHASE", ""),
        ]
        for col_idx, val in enumerate(data_vals, start=2):
            c        = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font   = DATA_FONT
            c.border = _hair_border()

        extra_keys = ["No of Days", "No of AMC GPs", "No of GPs Unknown previously UP"]
        for col_idx, key in zip([12, 13, 14], extra_keys):
            c      = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            c.font = EXTRA_FONT

    # Total row
    total_row = N_ROWS + 5
    ws.cell(row=total_row, column=11, value="Total").font = Font(name="Times New Roman", size=12)
    ws.cell(row=total_row, column=13, value=f"=SUM(M5:M{total_row-1})").font = EXTRA_FONT
    ws.cell(row=total_row, column=14, value=f"=SUM(N5:N{total_row-1})").font = EXTRA_FONT

    # Blank row + footer
    footer_row = total_row + 2
    ws.merge_cells(f"A{total_row+1}:N{total_row+1}")
    ws[f"A{total_row+1}"] = ""
    ws.merge_cells(f"A{footer_row}:N{footer_row}")
    ws[f"A{footer_row}"] = f"Report Generated at : {generated_at}- By User : {generated_by}"
    ws[f"A{footer_row}"].font = Font(name="Times New Roman", size=12)

    # Column widths
    COL_WIDTHS = {
        "A": 7,  "B": 14, "C": 21, "D": 16, "E": 20,
        "F": 22, "G": 15, "H": 14, "I": 12, "J": 22,
        "K": 10, "L": 16, "M": 15, "N": 32,
    }
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "A5"

    # ── 9. Return as bytes ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── CLI entry point ───────────────────────────────────────────────────────────

def main():
    """Run from the terminal: python olt_report.py"""
    base = Path(__file__).parent
    now  = datetime.now()

    permanent_path = _find_file(base, "permanent", "BA_OA_UP_WEST_LIST.xlsx")
    monthly_path   = _find_file(base, "monthly",   "AMC_GP_Status_June_26*.xlsx")
    daily_path     = _find_file(base, "daily",     "report_*.xlsx")
    daily_path2 = _find_file(base, "daily2",    "report_*.xlsx")

    print(f"[+] Permanent : {permanent_path.name}")
    print(f"[+] Monthly   : {monthly_path.name}")
    print(f"[+] Daily     : {daily_path.name}")
    print(f"[+] Daily2     : {daily_path2.name}")
    

    excel_bytes = generate_report(permanent_path, monthly_path, daily_path, daily_path2, now=now)

    out_path = base / "daily" / f"OLT_Down_Report_{now.strftime('%Y-%m-%d')}.xlsx"
    out_path.write_bytes(excel_bytes)
    print(f"\n✅  Report saved → {out_path}")


if __name__ == "__main__":
    main()