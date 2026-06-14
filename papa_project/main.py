"""
OLT Down Report Generator
=========================
Folder structure:
  permanent/   → BA_OA_UP_WEST_LIST.xlsx          (never changes)
  monthly/     → AMC_GP_Status_June_26_*.xlsx      (updated monthly)
  daily/       → report_<id>_<date>.xlsx           (downloaded every day)

Output: daily/OLT_Down_Report_<date>.xlsx
"""

import glob
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
NOW  = datetime.now()

# ── helpers ──────────────────────────────────────────────────────────────────

def find_file(folder: str, pattern: str) -> Path:
    matches = glob.glob(str(BASE / folder / pattern))
    if not matches:
        raise FileNotFoundError(f"No file matching '{pattern}' in {folder}/")
    return Path(sorted(matches)[-1])

HAIR = Side(border_style="hair", color="000000")
def hair_border(top=True, bottom=True):
    return Border(
        left=HAIR, right=HAIR,
        top=HAIR  if top    else Side(),
        bottom=HAIR if bottom else Side(),
    )

def age_bucket(raw_str: str) -> str:
    try:
        dt   = datetime.strptime(str(raw_str).strip(), "%d-%m-%Y %H:%M:%S")
        days = (NOW - dt).total_seconds() / 86400
        if days < 1:
            return "Less than 1 Day"
        elif days <= 3:
            return "More than 1 Day"
        elif days <= 7:
            return "More than 3 Days"
        else:
            return "More than 7 Days"
    except Exception:
        return "Unknown"


# ── 1. locate files ───────────────────────────────────────────────────────────

permanent_file = find_file("permanent", "BA_OA_UP_WEST_LIST.xlsx")
monthly_file   = find_file("monthly",   "AMC_GP_Status_June_26*.xlsx")
daily_file     = find_file("daily",     "report_*.xlsx")

print(f"[+] Permanent : {permanent_file.name}")
print(f"[+] Monthly   : {monthly_file.name}")
print(f"[+] Daily     : {daily_file.name}")


# ── 2. BA / OA mapping ───────────────────────────────────────────────────────

ba_oa = pd.read_excel(permanent_file, dtype=str)
ba_oa.columns = ba_oa.columns.str.strip()
ba_oa_map = (
    ba_oa[["BA", "OA", "BLOCK"]]
    .copy()
    .assign(BLOCK=lambda d: d["BLOCK"].str.strip().str.upper())
    .drop_duplicates(subset=["BLOCK"])
)


# ── 3. Daily report ───────────────────────────────────────────────────────────

report = pd.read_excel(daily_file, skiprows=3, dtype=str)
report.columns = report.columns.str.strip()
report.dropna(how="all", inplace=True)
report.reset_index(drop=True, inplace=True)
report["_BLOCK_KEY"] = report["BLOCK"].str.strip().str.upper()


# ── 4. Merge BA / OA before DISTRICT ─────────────────────────────────────────

merged = report.merge(
    ba_oa_map.rename(columns={"BLOCK": "_BLOCK_KEY"}),
    on="_BLOCK_KEY", how="left"
)
merged.drop(columns=["_BLOCK_KEY"], inplace=True)

# reorder: BA, OA just before DISTRICT
cols = list(merged.columns)
for c in ["BA", "OA"]:
    cols.remove(c)
di = cols.index("DISTRICT")
cols = cols[:di] + ["BA", "OA"] + cols[di:]
merged = merged[cols]


# ── 5. DOWN SINCE bucket ──────────────────────────────────────────────────────

merged["No of Days"] = merged["STATE CHANGE TIME"].apply(age_bucket)


# ── 6. AMC GP counts from monthly file ───────────────────────────────────────

xl_monthly = pd.ExcelFile(monthly_file)
olt_sheets = [s for s in xl_monthly.sheet_names if s.lower().startswith("olt")]
olt_sheet  = sorted(olt_sheets)[-1]
print(f"[+] AMC GP sheet: {olt_sheet}")

olt_raw = pd.read_excel(monthly_file, sheet_name=olt_sheet, skiprows=1, header=0)
olt_raw.columns = olt_raw.columns.str.strip()

amc_gp = (
    olt_raw[["OLT IP", "AMC GP Count"]]
    .dropna(subset=["OLT IP"])
    .copy()
    .assign(**{"OLT IP": lambda d: d["OLT IP"].astype(str).str.strip()})
)
amc_gp["AMC GP Count"] = pd.to_numeric(amc_gp["AMC GP Count"], errors="coerce")
amc_gp = amc_gp.groupby("OLT IP", as_index=False)["AMC GP Count"].sum()
amc_gp.rename(columns={"AMC GP Count": "No of AMC GPs"}, inplace=True)


# ── 7. Join AMC GP on BLOCK NODE IP ──────────────────────────────────────────

merged["_OLT_KEY"] = merged["BLOCK NODE IP"].astype(str).str.strip()
final = merged.merge(
    amc_gp.rename(columns={"OLT IP": "_OLT_KEY"}),
    on="_OLT_KEY", how="left"
)
final.drop(columns=["_OLT_KEY"], inplace=True)

# Fill NaN counts with 0
final["No of AMC GPs"] = pd.to_numeric(final["No of AMC GPs"], errors="coerce").fillna(0).astype(int)
final["No of GPs Unknown previously UP"] = ""


# ── 8. Select and order final columns (match reference exactly) ───────────────

OUTPUT_COLS = [
    "BA", "OA", "DISTRICT", "BLOCK", "BLOCK NODE LOCATION",
    "BLOCK NODE IP", "ALARM REASON", "VENDOR",
    "STATE CHANGE TIME", "PHASE",
    "No of Days", "No of AMC GPs", "No of Gps Unknown previously UP",
]
final = final[[c for c in OUTPUT_COLS if c in final.columns]]
final.reset_index(drop=True, inplace=True)
N_ROWS = len(final)


# ── 9. Build Excel output ─────────────────────────────────────────────────────

today_str = NOW.strftime("%d/%m/%Y")
today_fn  = NOW.strftime("%Y-%m-%d")
out_path  = BASE / "daily" / f"OLT_Down_Report_{today_fn}.xlsx"

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "OLT Down Report"

# ── row 1: title (merged A1:M1) ──
title_text = (
    f" Total NON_OPERATIONAL Report  STATE NAME : UTTAR PRADESH WEST"
    f"  Date : {today_str} ,Time : 10:00  and  Phase : BHARATNET"
)
ws.merge_cells("A1:M1")
ws["A1"] = title_text
ws["A1"].font      = Font(name="Courier New", bold=True, size=18)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 24

# ── row 2: record count (merged A2:M2) ──
generated_at = NOW.strftime("%d-%m-%Y %H:%M:%S")
ws.merge_cells("A2:M2")
ws["A2"] = f"Showing Records 1 to {N_ROWS}   , Report Generated at : {generated_at}"
ws["A2"].font      = Font(name="Times New Roman", size=12)
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 15.75

# ── row 3: blank (merged) ──
ws.merge_cells("A3:M3")
ws["A3"] = ""
ws.row_dimensions[3].height = 15.75

# ── row 4: header ──
HEADERS = [
    "Sr.No", "BA", "OA", "DISTRICT", "BLOCK", "BLOCK NODE LOCATION",
    "BLOCK NODE IP", "ALARM REASON", "VENDOR", "STATE CHANGE TIME",
    "PHASE", "No of Days", "No of AMC GPs", "No of Gps Unknown previously UP",
]
ws.row_dimensions[4].height = 131.25

for col_idx, hdr in enumerate(HEADERS, start=1):
    cell = ws.cell(row=4, column=col_idx, value=hdr)
    cell.font      = Font(name="Times New Roman", bold=True, size=14)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    # columns A–K get full hair border; L–N get left+right only (matches reference)
    if col_idx <= 11:
        cell.border = hair_border(top=True, bottom=True)
    else:
        cell.border = hair_border(top=False, bottom=False)

# ── rows 5+: data ──
DATA_FONT    = Font(name="Times New Roman", size=12)
EXTRA_FONT   = Font(name="Calibri", size=11)          # cols L–N

for row_idx, (_, row) in enumerate(final.iterrows(), start=5):
    sr_no = row_idx - 4
    # Sr.No
    c = ws.cell(row=row_idx, column=1, value=sr_no)
    c.font   = DATA_FONT
    c.border = hair_border()

    data_vals = [
        row.get("BA", ""), row.get("OA", ""), row.get("DISTRICT", ""),
        row.get("BLOCK", ""), row.get("BLOCK NODE LOCATION", ""),
        row.get("BLOCK NODE IP", ""), row.get("ALARM REASON", ""),
        row.get("VENDOR", ""), row.get("STATE CHANGE TIME", ""),
        row.get("PHASE", ""),
    ]
    for col_idx, val in enumerate(data_vals, start=2):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font   = DATA_FONT
        c.border = hair_border()

    # cols L, M, N — No of Days / AMC GPs / Unknown GPs
    for col_idx, key in zip([12, 13, 14], ["No of Days", "No of AMC GPs", "No of Gps Unknown previously UP"]):
        val = row.get(key, "")
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font   = EXTRA_FONT
        # no border (matches reference)

# ── Total row ──
total_row = N_ROWS + 5
ws.cell(row=total_row, column=11, value="Total").font = Font(name="Times New Roman", size=12)
ws.cell(row=total_row, column=13, value=f"=SUM(M5:M{total_row-1})").font = EXTRA_FONT
ws.cell(row=total_row, column=14, value=f"=SUM(N5:N{total_row-1})").font = EXTRA_FONT

# ── blank row then footer ──
footer_row = total_row + 2
ws.merge_cells(f"A{total_row+1}:M{total_row+1}")
ws[f"A{total_row+1}"] = ""

ws.merge_cells(f"A{footer_row}:M{footer_row}")
ws[f"A{footer_row}"] = f"Report Generated at : {generated_at}- By User : bsupwag2"
ws[f"A{footer_row}"].font = Font(name="Times New Roman", size=12)

# ── column widths (match reference where defined, sensible defaults elsewhere) ──
COL_WIDTHS = {
    "A": 7,    # Sr.No
    "B": 14,   # BA
    "C": 21,   # OA
    "D": 16,   # DISTRICT
    "E": 20,   # BLOCK
    "F": 22,   # BLOCK NODE LOCATION
    "G": 15,   # BLOCK NODE IP
    "H": 14,   # ALARM REASON
    "I": 12,   # VENDOR
    "J": 22,   # STATE CHANGE TIME
    "K": 10,   # PHASE
    "L": 16,   # No of Days
    "M": 15,   # No of AMC GPs
    "N": 32,   # No of Gps Unknown previously UP
}
for letter, width in COL_WIDTHS.items():
    ws.column_dimensions[letter].width = width

# freeze panes below header
ws.freeze_panes = "A5"

wb.save(out_path)
print(f"\n✅  Report saved → {out_path}")
print(f"    Rows : {N_ROWS}  |  Columns : {len(HEADERS)}")